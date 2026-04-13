import logging
import re
from pathlib import Path
from typing import Any, Dict, Optional

# Optional imports handled gracefully
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    openpyxl = None

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
except ImportError:
    SimpleDocTemplate = None

logger = logging.getLogger("BusinessPlanAgent.Exporters")

def _ensure_output_dir() -> str:
    """Ensure the output directory exists."""
    output_dir = Path("output")
    output_dir.mkdir(parents=True, exist_ok=True)
    return str(output_dir)

def clean_llm_text(text: str) -> str:
    """
    TASK 1: Clean Markdown artifacts and extra spaces.
    Removes: # headers, ** bold, and normalizes spacing.
    """
    if not text:
        return ""
    # 1. Remove Markdown Headers (###, ##, #)
    text = re.sub(r'#+\s*', '', text)
    # 2. Remove Bold Markers (**)
    text = text.replace('**', '')
    # 3. Normalize whitespace (keeps line structure for SWOT)
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    return '\n'.join(lines)

def clean_swot_data(text: str) -> Dict[str, list]:
    """
    TASK 3: Efficient SWOT parsing with structured cleaning.
    """
    categories = {
        "STRENGTHS": [], "WEAKNESSES": [], "OPPORTUNITIES": [], "THREATS": []
    }
    
    current_cat = None
    # Pre-clean the whole block
    text = clean_llm_text(text)
    
    for line in text.split('\n'):
        up = line.upper()
        if "STRENGTH" in up: current_cat = "STRENGTHS"
        elif "WEAKNESS" in up: current_cat = "WEAKNESSES"
        elif "OPPORTUN" in up: current_cat = "OPPORTUNITIES"
        elif "THREAT" in up: current_cat = "THREATS"
        elif current_cat:
            # Clean bullet points markers (1., -, *, etc)
            clean_point = re.sub(r'^([a-zA-Z0-9][\.\)\-]\s*|[\*\-\•\>\s]+)', '', line).strip()
            if clean_point:
                categories[current_cat].append(clean_point)
                
    return categories

def generate_excel(state: Dict[str, Any]) -> Dict[str, Any]:
    """
    TASK 4: Premium Excel structure with Borders and Zebra striping.
    """
    if not openpyxl:
        logger.error("Skipping Excel: openpyxl missing.")
        return state

    try:
        _ensure_output_dir()
        file_path = "output/business.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Analysis"
        
        # Colors & Fonts
        NAVY = "1F4E78"
        WHITE = "FFFFFF"
        STRIPE = "F2F2F2"
        
        header_font = Font(bold=True, color=WHITE, size=11)
        header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        stripe_fill = PatternFill(start_color=STRIPE, end_color=STRIPE, fill_type="solid")
        
        thin_side = Side(style='thin')
        all_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # 1. Headers
        headers = ["METRIC", "YEAR 1", "YEAR 2", "YEAR 3"]
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = all_border
            
        # 2. Financial Rows
        projections = state.get("projections", [])
        metrics = ["Revenue", "Profit", "ROI"]
        
        for r_idx, metric in enumerate(metrics, start=2):
            # Row label
            ws.cell(row=r_idx, column=1, value=metric).font = Font(bold=True)
            ws.cell(row=r_idx, column=1).border = all_border
            
            for c_idx, data in enumerate(projections, start=2):
                cell = ws.cell(row=r_idx, column=c_idx)
                val = data.get(metric.lower(), 0)
                cell.value = val
                cell.border = all_border
                
                # Format
                if metric == "ROI":
                    cell.number_format = '0.00%'
                else:
                    cell.number_format = '"$"#,##0.00'
                
                # Zebra Striping
                if r_idx % 2 == 1:
                    cell.fill = stripe_fill
                    ws.cell(row=r_idx, column=1).fill = stripe_fill

        # 3. Highlighted NPV Row
        npv_row = 5
        ws.cell(row=npv_row, column=1, value="NET PRESENT VALUE (NPV)").font = Font(bold=True)
        ws.cell(row=npv_row, column=2, value=state.get("npv", 0)).font = Font(bold=True)
        ws.cell(row=npv_row, column=2).number_format = '"$"#,##0.00'
        
        for c in range(1, 5):
            ws.cell(row=npv_row, column=c).border = all_border
            ws.cell(row=npv_row, column=c).alignment = Alignment(horizontal="center")

        ws.column_dimensions['A'].width = 30
        for col in ['B', 'C', 'D']:
            ws.column_dimensions[col].width = 18

        wb.save(file_path)
        logger.info(f"Polished Excel generated -> {file_path}")
    except Exception as e:
        logger.error(f"Excel error: {e}")
    return state

def generate_pdf(state: Dict[str, Any]) -> Dict[str, Any]:
    """
    TASK 2: Fixed PDF structure with high-end layout and NPV protection.
    """
    if not SimpleDocTemplate:
        logger.error("Skipping PDF: reportlab missing.")
        return state

    try:
        _ensure_output_dir()
        file_path = "output/business.pdf"
        doc = SimpleDocTemplate(file_path, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        
        # Professional Styles Extension
        styles.add(ParagraphStyle(name='ProjectTitle', fontSize=24, leading=28, textColor=colors.HexColor("#1F4E78"), fontName='Helvetica-Bold', spaceAfter=12))
        styles.add(ParagraphStyle(name='CategoryHeader', fontSize=12, leading=14, spaceBefore=15, spaceAfter=8, textColor=colors.HexColor("#2C3E50"), fontName='Helvetica-Bold'))
        styles.add(ParagraphStyle(name='TableLabel', fontSize=10, fontName='Helvetica-Bold', alignment=0))
        styles.add(ParagraphStyle(name='TableCell', fontSize=10, fontName='Helvetica', alignment=2))
        styles.add(ParagraphStyle(name='TableHead', fontSize=10, fontName='Helvetica-Bold', alignment=1, textColor=colors.whitesmoke))
        
        elements = []
        
        # 1. Header Section
        elements.append(Paragraph("INDUSTRIE IA : STRATEGIC ANALYSIS", styles['ProjectTitle']))
        elements.append(Paragraph(f"Analysis For: {state.get('specs', {}).get('type', 'Industrial Enterprise')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # 2. 3-Year Financial Table
        elements.append(Paragraph("FINANCIAL GROWTH PROJECTIONS", styles['CategoryHeader']))
        
        projections = state.get("projections", [])
        # Correctly formatted headers using Paragraph objects (No raw HTML visible)
        table_data = [
            [Paragraph("METRIC", styles['TableHead']), 
             Paragraph("YEAR 1", styles['TableHead']), 
             Paragraph("YEAR 2", styles['TableHead']), 
             Paragraph("YEAR 3", styles['TableHead'])]
        ]
        
        for m in ["Revenue", "Profit", "ROI"]:
            row = [Paragraph(f"<b>{m}</b>", styles['Normal'])]
            for data in projections:
                val = data.get(m.lower(), 0)
                text = f"{val:.1%}" if m == "ROI" else f"${val:,.0f}"
                row.append(Paragraph(text, styles['TableCell']))
            table_data.append(row)
            
        t_financials = Table(table_data, colWidths=[140, 110, 110, 110])
        t_financials.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1F4E78")),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.HexColor("#F2F2F2")]),
        ]))
        elements.append(t_financials)
        elements.append(Spacer(1, 25))
        
        # 3. NPV Section (TASK 2.A: Strict Currency Formatting)
        elements.append(Paragraph("NET PRESENT VALUE (NPV) & VALUE ASSESSMENT", styles['CategoryHeader']))
        npv_val = state.get("npv", 0)
        formatted_npv = f"${npv_val:,.2f}" # NEVER percentage
        
        decision = state.get("decision", "N/A")
        
        summary_data = [
            [Paragraph("<b>ESTIMATED PROJECT NPV:</b>", styles['Normal']), Paragraph(f"<b>{formatted_npv}</b>", styles['Normal'])],
            [Paragraph("<b>RECOMMENDATION status:</b>", styles['Normal']), Paragraph(f"<b>{decision}</b>", styles['Normal'])]
        ]
        t_summary = Table(summary_data, colWidths=[200, 270])
        t_summary.setStyle(TableStyle([
            ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('TOPPADDING', (0,0), (-1,-1), 10),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ]))
        elements.append(t_summary)
        elements.append(Spacer(1, 25))
        
        # 4. SWOT Section (TASK 3: Structured categories)
        elements.append(Paragraph("INDUSTRIAL SWOT ANALYSIS", styles['CategoryHeader']))
        swot_txt = state.get("swot", "")
        swot_categories = clean_swot_data(swot_txt)
        
        for category, points in swot_categories.items():
            elements.append(Paragraph(category, styles['TableLabel']))
            if not points:
                elements.append(Paragraph("&bull; Data not available.", styles['Normal']))
            else:
                for pt in points:
                    # Clean the point text specifically before bullet
                    elements.append(Paragraph(f"&bull; {pt}", styles['Normal']))
            elements.append(Spacer(1, 8))
            
        doc.build(elements)
        logger.info(f"Corrected PDF generated -> {file_path}")
    except Exception as e:
        logger.error(f"PDF error: {e}", exc_info=True)
    return state
